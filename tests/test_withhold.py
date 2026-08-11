"""
When the model refuses to value a name, the whole page must refuse (offline, no network).

    python tests/test_withhold.py

THE BUG. KSPI, 2026-08-05, signed out. The headline read "Fair value: Not DCF-valuable,
Upside: n/a" with the guard's reason under it — and then FAIR VALUE & SCENARIOS printed
BEAR $620.31 (+573%) / BASE $1,289.68 (+1299%) / BULL $2,888.33 (+3033%): the exact number
the guard had just withheld, three inches below the notice withholding it, in green.

Six more cards did the same thing (Monte Carlo median and "100% of trials above the price",
the sensitivity grid, the comps implied values, the reverse-DCF verdict, the FCF projection,
and a 93/100 "Strong Buy" gauge). The fix is a rule, not seven patches: nothing computed
from a withheld fair value is published, and the numbers are stripped from the API response
rather than merely not drawn — so "it isn't on the page" is checkable here, offline, instead
of by squinting at a browser.

The fixture below is the REAL KSPI payload shape with the REAL figures that shipped, so a
regression reproduces the actual bug rather than a sanitised version of it.
"""
import json
import os
import re
import sys
import types

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from valuation.web import withhold                      # noqa: E402

APP_JS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                      "valuation", "web", "static", "app.js")

# ===========================================================================================
#  known_failure — the same mechanism `tests/test_guards.py` uses, for the same reason.
#
#  A test that encodes a bug owned by ANOTHER lane cannot be allowed to turn this suite red:
#  the gate auto-merges to main, so a red suite here would block every unrelated lane. But
#  landing the test only once the fix arrives means nobody ever sees it fail, and the leak
#  goes back to being found by a production probe. XFAIL is the compromise — the bug is
#  encoded, executable, and visible on every run, and it flips to XPASS (printed loudly) the
#  day the owning lane fixes it.
# ===========================================================================================
_KNOWN_FAILURES = {}


def known_failure(reason: str, lane: str):
    def deco(fn):
        _KNOWN_FAILURES[fn.__name__] = (reason, lane)
        return fn
    return deco

PRICE = 92.19
REASON = ("Cannot value this name: the model's $1,289.60 is 14.0x the $92.19 price. That gap "
          "is a data problem (currency or share count), not an opportunity, so no fair value "
          "is published.")

#: Every figure that rendered on the live page for KSPI while the headline said the name
#: could not be valued. Each one must be gone after withholding.
LEAKED = [620.2688355663354, 1289.5983215433105, 2888.146436947473,   # scenario cards
          2293.6874204966884, 928.6315450198296, 5839.11241046628,    # raw DCF cone
          2335.725200478383, 872.8932813115293, 6340.114354833571,    # Monte Carlo
          2798.85, 3448.83, 4746.12, 8632.67,                         # sensitivity grid
          326.3166344204799, 360.1179, 292.8437,                      # comps implied
          987.2461624629182]                                          # growth lens


def _payload(valuable=False):
    """The /api/value payload for KSPI, as the engine produced it."""
    return {
        "company": {"ticker": "KSPI", "name": "Kaspi.kz", "price": PRICE,
                    "currency": "USD", "financial_currency": "KZT"},
        "classification": {"regime": "growth", "dcf_reliability": "low"},
        "wacc": {"wacc": 0.0503},
        "assumptions": {"start_growth": 0.3428, "notes": []},
        "scenarios": {
            "bear": {"per_share": 928.6315450198296, "equity_value": 1.7e8, "rows": [{"year": 1, "revenue": 11496.2}]},
            "base": {"per_share": 2293.6874204966884, "equity_value": 4.3e8, "rows": [{"year": 1, "revenue": 11496.2}]},
            "bull": {"per_share": 5839.11241046628, "equity_value": 1.1e9, "rows": [{"year": 1, "revenue": 11496.2}]},
            "bear_price": 928.6315450198296, "base_price": 2293.6874204966884,
            "bull_price": 5839.11241046628},
        "montecarlo": {"trials": 2000, "mean": 2400.1, "median": 2335.725200478383, "std": 900.0,
                       "p5": 700.0, "p10": 872.8932813115293, "p25": 1400.0, "p75": 4000.0,
                       "p90": 6340.114354833571, "p95": 7100.0, "prob_undervalued": 1.0,
                       "price": PRICE, "hist_bins": [400.0, 900.0], "hist_counts": [12, 44]},
        "reverse": {"implied_avg_growth": -0.09795, "base_avg_growth": 0.20205,
                    "growth_verdict": "Market prices in only ~-9.8% avg growth — below our "
                                      "base (20.2%). Expectations look cheap.",
                    "margin_verdict": "…or only a ~20.3% terminal margin vs our 40.3%."},
        "comps": {"subject": {"pe": 7.67998452136027, "ps": 2.046261889466831,
                              "ev_sales": 1.6520418341714658, "ev_ebitda": 5.054346999453748},
                  "benchmark": {"pe": 30, "ev_ebitda": 22, "ps": 6.5, "ev_sales": 6.5},
                  "benchmark_source": "Technology sector benchmark",
                  "implied": {"pe": 360.1179, "ps": 292.8437, "ev_sales": 310.6045,
                              "ev_ebitda": 341.7281},
                  "comps_fair_value": 326.3166344204799},
        "sensitivity": {"wacc_axis": [0.03, 0.04], "growth_axis": [0.02, 0.03],
                        "grid": [[2798.85, 3448.83, 4746.12, 8632.67, None]],
                        "base_wacc": 0.0503, "base_growth": 0.025},
        "score": {"score": 93, "recommendation": "Strong Buy", "confidence": "medium",
                  "subscores": {"valuation": 100.0, "quality": 90.5, "growth": 87.5,
                                "health": 100.0, "momentum": 78.0},
                  "weights": {"valuation": 0.3, "quality": 0.2, "growth": 0.2,
                              "health": 0.15, "momentum": 0.15},
                  "drivers": ["Monte Carlo: 100% of trials value it above the price.",
                              "Comps imply $326.32 (+254%).",
                              "ROIC 24% vs WACC 5% → +19% value-creation spread.",
                              "Forward revenue growth ~34%.",
                              "Net debt/EBITDA -0.5x, positive FCF."]},
        "ai": None,
        "fair_value_blend": {
            "value": None if not valuable else 1289.5983215433105,
            "valuable": valuable, "reason": "" if valuable else REASON,
            "method": "39% DCF · 32% multiples · 28% growth (revenue multiple)",
            "confidence": "low", "value_low": 620.2688355663354, "value_high": 2888.146436947473,
            "lenses": {"dcf": {"value": 2293.7359, "weight": 0.395},
                       "multiples": {"value": 326.3236, "weight": 0.323},
                       "growth": {"value": 987.2462, "weight": 0.282}},
            "notes": []},
        "growth_lens": {"value": 987.2461624629182, "applies": True,
                        "revenue_at_horizon": 19188.6, "exit_multiple": 11.026},
        "fair_value_scenarios": {"method": "39% DCF · 32% multiples · 28% growth",
                                 "bear": 620.2688355663354, "base": 1289.5983215433105,
                                 "bull": 2888.146436947473},
        "base_fair_value": None if not valuable else 1289.5983215433105,
        "dcf_per_share": 2293.6874204966884,
        "upside": None,
        "warnings": [REASON],
        "sources": ["FMP"],
    }


def _numbers(obj, path="", out=None):
    """Every numeric leaf in a structure, with the path it sits at."""
    out = [] if out is None else out
    if isinstance(obj, dict):
        for k, v in obj.items():
            _numbers(v, f"{path}.{k}", out)
    elif isinstance(obj, (list, tuple)):
        for i, v in enumerate(obj):
            _numbers(v, f"{path}[{i}]", out)
    elif isinstance(obj, bool) or obj is None:
        pass
    elif isinstance(obj, (int, float)):
        out.append((path, float(obj)))
    return out


# --------------------------------------------------------------------------- #
# The rule itself
# --------------------------------------------------------------------------- #
def test_a_valuable_name_is_returned_completely_untouched():
    """The withholding must be invisible on every name that was not refused — otherwise the
    cure is worse than the bug."""
    ok = _payload(valuable=True)
    assert withhold.is_withheld(ok) is False
    out = withhold.withhold_derived_figures(ok)
    assert out is ok, "a publishable name must not be copied, altered or marked"
    assert json.dumps(out, sort_keys=True) == json.dumps(_payload(valuable=True), sort_keys=True)


def test_the_guard_firing_is_what_triggers_withholding():
    """Both halves of the renderer's own test (app.js: `base_fair_value == null &&
    valuable === false`). A missing fair value alone is a data gap, not a refusal."""
    assert withhold.is_withheld(_payload()) is True
    gap = _payload(valuable=True)
    gap["base_fair_value"] = None                      # no blend refusal — just no number
    gap["fair_value_blend"]["value"] = None
    assert withhold.is_withheld(gap) is False
    assert withhold.is_withheld({}) is False
    assert withhold.is_withheld(None) is False


def test_the_scenario_cone_the_bug_was_reported_for_is_gone():
    out = withhold.withhold_derived_figures(_payload())
    fvs = out["fair_value_scenarios"]
    assert fvs["bear"] is None and fvs["base"] is None and fvs["bull"] is None
    assert fvs["withheld"] is True
    # the method label is not a number and stays, so the card can still say what it WOULD
    # have been valued as
    assert "DCF" in fvs["method"]
    for k in ("bear_price", "base_price", "bull_price"):
        assert out["scenarios"][k] is None
    for case in ("bear", "base", "bull"):
        assert out["scenarios"][case]["per_share"] is None
        assert out["scenarios"][case]["rows"] == []
    assert out["dcf_per_share"] is None
    assert out["fair_value_blend"]["value_low"] is None
    assert out["fair_value_blend"]["value_high"] is None
    assert out["fair_value_blend"]["lenses"] == {}
    assert out["growth_lens"] is None


def test_the_same_valuation_in_its_other_costumes_is_gone_too():
    """Monte Carlo, the sensitivity grid and the reverse-DCF read are all the withheld
    number wearing something else. `prob_undervalued` matters most: it is the share of
    trials of the withheld DCF that beat the price, and it printed as "100% of trials value
    it above today's price"."""
    out = withhold.withhold_derived_figures(_payload())
    mc = out["montecarlo"]
    for k in ("mean", "median", "p5", "p10", "p25", "p75", "p90", "p95",
              "prob_undervalued", "hist_bins", "hist_counts"):
        assert mc[k] is None, f"montecarlo.{k} survived"
    assert mc["withheld"] is True and mc["price"] == PRICE   # the price is not our output
    assert out["sensitivity"]["grid"] == []
    assert out["sensitivity"]["withheld"] is True
    assert out["reverse"]["growth_verdict"] is None
    assert out["reverse"]["implied_avg_growth"] is None


def test_comps_keep_the_ratios_and_lose_the_implied_dollars():
    """A multiple is currency-neutral; the per-share value implied by it is not, which is
    how a $92 stock showed a $326 implied value. The distinction is the point of the card."""
    out = withhold.withhold_derived_figures(_payload())
    assert out["comps"]["subject"]["pe"] == 7.67998452136027
    assert out["comps"]["benchmark"]["pe"] == 30
    assert out["comps"]["implied"] == {}
    assert out["comps"]["comps_fair_value"] is None
    assert out["comps"]["withheld"] is True


def test_no_withheld_figure_survives_anywhere_in_the_valuation_blocks():
    """The catch-all. Rather than listing the seven cards known to have leaked, walk every
    number in every valuation block and require it to be plausible against the price — the
    guard's own >5x band. A new card that starts republishing the DCF fails here without
    anyone remembering to add it to a list."""
    out = withhold.withhold_derived_figures(_payload())
    blocks = ("scenarios", "montecarlo", "sensitivity", "comps", "reverse",
              "fair_value_blend", "growth_lens", "fair_value_scenarios", "score")
    counts = (".trials",)          # a trial COUNT is not a value figure
    for b in blocks:
        for path, v in _numbers(out.get(b), b):
            if path.endswith(counts):
                continue
            assert abs(v) <= PRICE * 5, f"{path} = {v} survived on a {PRICE} stock"
    for key in ("base_fair_value", "dcf_per_share", "upside"):
        assert out[key] is None
    # and none of the exact figures that shipped are anywhere in the response body, except
    # the guard's reason, which quotes one on purpose (see below)
    scrubbed = {k: v for k, v in out.items() if k not in ("warnings", "withheld")}
    scrubbed["fair_value_blend"] = {k: v for k, v in out["fair_value_blend"].items()
                                    if k != "reason"}
    body = json.dumps(scrubbed)
    for v in LEAKED:
        assert f"{v}" not in body, f"{v} is still on the wire"
        assert f"{v:,.2f}" not in body


def test_the_reason_keeps_the_number_because_that_is_the_evidence():
    """The one permitted place. "the model's $1,289.60 is 14.0x the $92.19 price" is the
    argument for withholding, not a valuation — deleting it would leave a refusal with no
    stated cause, which is worse."""
    out = withhold.withhold_derived_figures(_payload())
    assert out["fair_value_blend"]["reason"] == REASON
    assert "$1,289.60" in out["withheld"]["reason"]
    assert out["withheld"]["cards"]["scenarios"]
    assert out["withheld"]["score_note"]


# --------------------------------------------------------------------------- #
# The score — the second costume, and the more serious one
# --------------------------------------------------------------------------- #
def test_the_score_is_published_as_partial_not_suppressed():
    """The history matters, because the right answer changed twice.

    `compute_score` used to be handed `base_fv=None`, which dropped only the margin-of-safety
    term (0.55) while `mc.prob_undervalued` (0.30, = 1.00 on KSPI, trials OF THE WITHHELD DCF)
    and `comps.comps_fair_value` (0.15) rebuilt it — the valuation sub-score printed 100.0/100
    on a name the model had declined to value, and the composite read 93 "Strong Buy". This
    page then refused to show the score at all, which was right while the number was
    contaminated.

    The engine fixed it (2026-08-06): the ENTIRE valuation sub-score is dropped and the >5x
    cap now falls back to `blend.withheld_value` instead of being dead. KSPI 93 -> 50. So the
    number underneath is sound, "Not rated" became an understatement, and what is published
    now is a PARTIAL score — marked as one everywhere it appears.
    """
    out = withhold.withhold_derived_figures(_payload())
    s = out["score"]
    assert s["score"] == 93, "the partial score itself must survive — it is not contaminated"
    assert s["recommendation"], "a partial score still carries its band label"
    assert s["partial"] is True
    assert s["subscores"]["valuation"] is None
    assert sorted(s["partial_of"]) == ["growth", "health", "momentum", "quality"]
    assert s["confidence"] == "low"
    # the components with no fair value in them survive — that is what makes a partial score
    # meaningful rather than a stub
    assert s["subscores"]["quality"] == 90.5 and s["subscores"]["momentum"] == 78.0


def test_the_engines_own_explanation_survives_this_filter():
    """A REGRESSION THIS FILE HAS ALREADY CAUSED ONCE IN DESIGN. The driver filter matched on
    markers alone, and the two drivers a withheld name now legitimately carries are

        "Valuation withheld — no fair-value, Monte Carlo or comps term contributes ..."
        "⚠ Model fair value is 11.3× the price — implausible ... Capped and flagged ..."

    A marker-only match deletes BOTH: the explanation the page is required to show, and the
    flag saying the number was capped. Every leaking driver stated a dollar amount; neither of
    these does. So the "$" is load-bearing, and this test is the reason it stays.
    """
    p = _payload()
    p["score"]["drivers"] = [
        "⚠ Model fair value is 11.3× the price — implausible; likely a data issue (currency, "
        "share count, or a one-off). Capped and flagged unreliable, not a recommendation.",
        "Valuation withheld — no fair-value, Monte Carlo or comps term contributes to this "
        "score. Scored on quality, growth, financial health and momentum only.",
        "Base fair value $1,289.60 vs $92.19 → +1299% margin of safety.",
        "Monte Carlo: 100% of trials value it above the price.",
        "Comps imply $326.32 (+254%).",
        "ROIC 36% vs WACC 5% → +31% value-creation spread.",
    ]
    drivers = withhold.withhold_derived_figures(p)["score"]["drivers"]
    joined = " ".join(drivers)
    assert "Valuation withheld" in joined, "the engine's explanation was filtered out"
    assert "Capped and flagged" in joined, "the cap flag was filtered out"
    assert "ROIC" in joined
    assert "$1,289.60" not in joined and "$326.32" not in joined
    assert "of trials value it above the price" not in joined
    assert len(drivers) == 3


def test_the_score_note_says_partial_and_says_it_is_not_comparable():
    """The greeks lane could not evaluate, on five names, whether a partial score and a full
    score mean the same thing at the same number. The labelling IS the mitigation, so the
    wording has to carry it rather than imply it."""
    note = withhold.SCORE_NOTE.lower()
    assert "partial" in note
    assert "not comparable" in note, "the one caveat that was explicitly routed here"
    assert "contributes nothing" in note or "withheld" in note
    assert "quality" in note and "momentum" in note


# --------------------------------------------------------------------------- #
# The renderer, checked at the source
# --------------------------------------------------------------------------- #
def _render_body():
    src = open(APP_JS, encoding="utf-8").read()
    i = src.index("function render(d)")
    return src[i:src.index("\nfunction metric(", i)], src


def test_the_renderer_draws_nothing_derived_from_a_withheld_value():
    """The server strips the figures, but the browser is the surface the reader sees, so it
    refuses independently: every call that draws a DCF-derived figure sits in the ELSE branch
    of `notValuable`. Two locks, because this bug was one lock failing."""
    body, _ = _render_body()
    assert "if (notValuable) {" in body and "withheldCards(d);" in body
    guarded = body.index("if (notValuable) {")
    els = body.index("} else {", guarded)
    for call in ("rangebar(", "scenarioCards(", "fcfChart(", "mcChart(", "sensBox(",
                 "reverseBox("):
        at = body.index(call)
        assert at > els, f"{call} is called outside the notValuable else-branch"
        assert body.count(call) == 1, f"{call} is called more than once in render()"


def test_the_stale_chart_path_is_closed():
    """Skipping a chart draw is not enough: Chart.js keeps the previous ticker's canvas, so
    a withheld name would have shown the LAST name's cone. `withheldCards` destroys both."""
    _, src = _render_body()
    i = src.index("function withheldCards(")
    fn = src[i:src.index("\n}", i)]
    assert 'killChart("fcf")' in fn and 'killChart("mc")' in fn
    for el in ("rangebar", "scenarioCards", "mcNote", "sensBox", "reverseBox", "fcfNote"):
        assert el in fn, f"withheldCards leaves #{el} holding whatever was there"


def test_the_refusal_says_why_rather_than_going_blank():
    """A blank card reads as "still loading" or "no data for this one". The reader is owed
    the same sentence the headline gave them, on every card that disappeared."""
    _, src = _render_body()
    assert "function withheldBox(" in src
    assert "Not published for this name." in src
    assert "Not rated." in src, "the gauge must say it is not rated, not just vanish"
    for key in ("scenarios", "montecarlo", "sensitivity", "fcf", "comps", "reverse", "score"):
        assert key in withhold.CARD_REASONS
        # the client carries the same copy as a fallback, so a card cannot render an empty
        # explanation if the server block is ever missing
        assert re.search(rf"\b{key}:\s*\"", src), f"no client fallback reason for {key}"


def test_the_gauge_marks_a_partial_score_on_the_dial_itself():
    """Not in a tooltip, not in a footnote. The greeks lane could not test whether a partial
    50 and a full 50 mean the same thing, so the label is the mitigation — it has to be
    unmissable, and it has to be impossible for the withheld branch to fall through into the
    full-gauge markup."""
    _, src = _render_body()
    i = src.index("function gauge(")
    fn = src[i:src.index("\n\n/* ---------- how the fair value", i)]
    assert "if (s == null)" in fn, "a genuinely absent score still refuses"
    j = fn.index("if (notValuable) {")
    partial = fn[j:fn.index("return;", j)]
    for mark in ("PARTIAL", "4 of 5 components", "Valuation withheld", "partial"):
        assert mark in partial, f"the partial dial never says {mark!r}"
    assert "score-num" in partial, "the number itself is still shown — it is not contaminated"
    assert fn.index("if (notValuable) {") < fn.index('class="rec" style="color:${col}'), \
        "the withheld branch must return before the full-gauge markup"


# --------------------------------------------------------------------------- #
# End to end, through the actual route
# --------------------------------------------------------------------------- #
def _stub_result(payload):
    blend = types.SimpleNamespace(valuable=payload["fair_value_blend"]["valuable"],
                                  reason=payload["fair_value_blend"]["reason"])
    return types.SimpleNamespace(to_dict=lambda: payload,
                                 base_fair_value=payload["base_fair_value"],
                                 fair_value_blend=blend, ai=None)


def _client_with_stub(payload):
    from valuation.web import app as webapp
    orig = webapp.value_ticker
    webapp.value_ticker = lambda *a, **k: _stub_result(payload)
    webapp.app.config["TESTING"] = True
    return webapp, orig


def test_the_api_response_itself_carries_no_withheld_figure():
    """The end the reader can actually inspect: view-source and the network tab. Anything
    still in this response is one console line away from being republished."""
    payload = _payload()
    webapp, orig = _client_with_stub(payload)
    try:
        with webapp.app.test_client() as c:
            body = c.post("/api/value", json={"ticker": "KSPI"}).get_data(as_text=True)
    finally:
        webapp.value_ticker = orig
    for v in LEAKED:
        assert str(v) not in body, f"{v} reached the browser"
    assert "1289.5983215433105" not in body
    assert "$1,289.60" in body, "the reason, which quotes it deliberately, must survive"
    compact = re.sub(r"\s+", "", body)
    # The score is PUBLISHED now (the engine no longer contaminates it) but must arrive
    # flagged partial, with the valuation component null.
    assert '"partial":true' in compact and '"valuation":null' in compact
    # the misleading data-gap message must not be attached to a deliberate refusal
    assert "Check the ticker symbol" not in body


def test_a_publishable_name_still_gets_its_numbers_through_the_route():
    payload = _payload(valuable=True)
    webapp, orig = _client_with_stub(payload)
    try:
        with webapp.app.test_client() as c:
            body = c.post("/api/value", json={"ticker": "NKE"}).get_data(as_text=True)
    finally:
        webapp.value_ticker = orig
    assert "1289.5983215433105" in body and "2293.6874204966884" in body


# --------------------------------------------------------------------------- #
# THE LIST SURFACES — a different code path, the same claim
#
# Everything above protects `/api/value` and the documents built from it. `/api/hotstocks` is
# PUBLIC and reaches a fair value through `estimate_fair_values` instead, which has no
# ceiling: its EV bridge is `3 + 2 x (net debt / market cap)` times the price. Measured on the
# real production snapshot (2026-08-06): 499 rows carried a fair value and one — AEG at 5.25x,
# $49.91 against $9.50, "blended / medium" — cleared the band the valuation page enforces.
# --------------------------------------------------------------------------- #
BAND = 5.0


def _row(ticker, price, fv, **kw):
    r = {"ticker": ticker, "price": price, "fair_value": fv, "market_cap": 1e9,
         "hot_score": 50.0, "sector": "Utilities", "extra": {}}
    if fv is not None and price:
        r["upside"] = fv / price - 1.0
    r.update(kw)
    return r


def test_the_row_guard_uses_the_valuation_pages_own_band_not_its_own_number():
    """One number, one meaning. The band is imported from `engine.pipeline.FV_BAND_HIGH`
    rather than restated here, so the list and the page cannot drift into two different
    definitions of "implausible" — which is exactly how this leak opened."""
    from valuation.engine.pipeline import FV_BAND_HIGH
    assert withhold._band() == float(FV_BAND_HIGH) == BAND


def test_an_implausible_row_loses_its_fair_value_and_gains_a_reason():
    rows = [_row("FINE", 10.0, 30.0),                    # 3.0x — published
            _row("EDGE", 10.0, 50.0),                    # exactly 5.0x — published
            _row("AEG", 9.50, 49.91),                    # 5.25x — the real one
            _row("LEV", 10.0, 330.0)]                    # 33x — the constructed one
    n = withhold.withhold_implausible_fair_values(rows)
    assert n == 2, "the band is > band, not >= band"
    assert rows[0]["fair_value"] == 30.0 and rows[1]["fair_value"] == 50.0
    for r in rows[2:]:
        assert r["fair_value"] is None and r["upside"] is None
        assert r["fair_value_withheld"] is True
        why = r["fair_value_withheld_reason"]
        # "Say why, don't just blank it" — a silently missing cell reads as a data gap and
        # invites someone to fill it back in.
        assert "5x band" in why or "5x" in why
        assert "no fair value is published" in why.lower()
    assert f"{49.91 / 9.50:.1f}x" in rows[2]["fair_value_withheld_reason"], \
        "the reason states the actual ratio, not a generic sentence"


def test_a_row_already_marked_withheld_is_honoured_even_below_the_band():
    """Fail-closed, and forward-compatible with the screener fix. Today
    `screen.py::_enrich_with_dcf` writes `fair_value = None` when the publication guard
    REFUSES a name, and `estimate_fair_values` then reads that as "no DCF yet" and substitutes
    a peer estimate — so KSPI, which the valuation page refuses outright, is served on the
    public hot list at $299.16 (3.24x). That erasure is the screener lane's to fix; when it
    starts marking those rows, this surface already refuses them."""
    rows = [_row("KSPI", 92.19, 299.16, fair_value_withheld=True)]
    assert withhold.withhold_implausible_fair_values(rows) == 1
    assert rows[0]["fair_value"] is None and rows[0]["upside"] is None


def test_the_guard_survives_junk_rows_without_throwing():
    rows = [{}, {"price": 0, "fair_value": 5.0}, {"price": None, "fair_value": None},
            {"price": "n/a", "fair_value": "x"}, _row("OK", 10.0, 12.0)]
    withhold.withhold_implausible_fair_values(rows)      # must not raise
    assert rows[-1]["fair_value"] == 12.0


class _FakeStore:
    """The minimum surface `/api/hotstocks` and `/api/whatdo` read."""

    DATE = "2026-08-06"

    def __init__(self, rows):
        self._rows = rows

    def latest_scan_date(self):
        return self.DATE

    def load_snapshot(self, date, top=None):
        import copy
        rows = copy.deepcopy(self._rows)
        return rows[:top] if top else rows

    def list_scans(self):
        return [{"scan_date": self.DATE, "universe_size": len(self._rows),
                 "scored": len(self._rows), "params": "{}"}]

    def get_meta(self, *a, **k):
        return {}


def _public_rows():
    """A snapshot shaped like the real one, including the leak. The leveraged row is what
    `estimate_fair_values` produces for net debt >> market cap; it is written directly here so
    the test does not depend on the screener lane's arithmetic staying put."""
    rows = [_row(f"N{i}", 10.0 + i, (10.0 + i) * 1.4) for i in range(8)]
    rows.append(_row("AEG", 9.50, 49.91, fair_value_method="blended",
                     fair_value_confidence="medium"))
    rows.append(_row("LEV", 10.0, 330.0, fair_value_method="multiples",
                     fair_value_confidence="medium"))
    return rows


def _walk_fair_values(obj, path="", out=None):
    """Every fair-value-bearing object anywhere in a response body.

    Carries the two DESCRIBING labels as well as the value (LA10, 2026-08-10), because the
    band walk below is blind to them by construction: a withheld row's ratio is `None`, so
    every ratio assertion passes while `fair_value_method: "blended"` ships beside a null.
    """
    out = [] if out is None else out
    if isinstance(obj, dict):
        if "fair_value" in obj and "price" in obj:
            out.append({"path": path, "fv": obj.get("fair_value"), "px": obj.get("price"),
                        "marked": obj.get("fair_value_withheld"),
                        "method": obj.get("fair_value_method"),
                        "confidence": obj.get("fair_value_confidence")})
        for k, v in obj.items():
            _walk_fair_values(v, f"{path}.{k}", out)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            _walk_fair_values(v, f"{path}[{i}]", out)
    return out


def test_no_public_api_response_carries_a_fair_value_past_the_band():
    """THE DURABLE PART. The Session 12 catch-all walked every number in `/api/value`; this
    walks every fair value in every PUBLICLY SERVED ROW, which is the walk that would have
    caught this leak. A new public list surface fails here the day it starts serving a
    `fair_value` next to a `price`, without anyone remembering to add it to a list."""
    from valuation.web import app as webapp
    orig = webapp._store
    webapp._store = lambda: _FakeStore(_public_rows())
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            for url in ("/api/hotstocks?top=100", "/api/whatdo?ticker=AEG",
                        "/api/whatdo?ticker=LEV", "/api/whatdo?ticker=N3"):
                body = c.get(url).get_json()
                pairs = _walk_fair_values(body, url)
                assert pairs, f"{url} served no fair-value rows — the test proved nothing"
                for row in pairs:
                    fv, px = row["fv"], row["px"]
                    if fv is None:
                        continue
                    assert px and fv / px <= BAND, \
                        f"{row['path']} published {fv} against a price of {px} ({fv / px:.1f}x)"
    finally:
        webapp._store = orig


def test_no_public_api_response_describes_a_fair_value_it_withheld():
    """LA10. The other half of the same walk, and the half the band cannot see.

    A withheld row has no ratio, so every assertion in the test above passes on it while the
    row still ships `fair_value_method: "blended"` and `fair_value_confidence: "medium"` —
    stating the method and the confidence of a number that is not in the payload. Same shape
    as the original KSPI bug (a figure surviving its own suppression), one level up: here it
    is the LABEL that survived the value.

    Non-vacuous by construction: the assertion at the end fails if the fixture stops producing
    withheld rows, so this cannot quietly become a test of nothing.
    """
    from valuation.engine.publication import ROW_WITHHELD_METHOD
    from valuation.web import app as webapp
    orig = webapp._store
    webapp._store = lambda: _FakeStore(_public_rows())
    webapp.app.config["TESTING"] = True
    seen_withheld = 0
    try:
        with webapp.app.test_client() as c:
            for url in ("/api/hotstocks?top=100", "/api/whatdo?ticker=AEG",
                        "/api/whatdo?ticker=LEV", "/api/whatdo?ticker=N3"):
                for row in _walk_fair_values(c.get(url).get_json(), url):
                    if not row["marked"]:
                        continue
                    seen_withheld += 1
                    p = row["path"]
                    assert row["fv"] is None, f"{p} is marked withheld and still carries a value"
                    assert row["method"] in (None, ROW_WITHHELD_METHOD), (
                        f"{p} is withheld and still describes the method as {row['method']!r} — "
                        f"a label that outlived the value it described")
                    assert row["confidence"] is None, (
                        f"{p} is withheld and still claims {row['confidence']!r} confidence in "
                        f"a number it did not publish")
    finally:
        webapp._store = orig
    assert seen_withheld >= 2, (
        f"only {seen_withheld} withheld rows in the walk — the fixture stopped producing the "
        f"case this test exists for, so it proved nothing")


# FIXED 2026-08-07 and the marker is REMOVED, which is the whole point of the mechanism.
#
# This was a `known_failure` from Session 17: `save_snapshot` wrote a FIXED 18-column INSERT
# with no column for `fair_value_withheld`, so a refusal recorded during the scan was thrown
# away when the row was persisted, and `estimate_fair_values` then read the surviving
# `fair_value=None` as "no DCF yet" and substituted a peer estimate. Measured on production
# 2026-08-06: KSPI served at $299.155 (3.24x) while its valuation page refused it at 11.2x.
#
# The greeks lane fixed it in `valuation/screener/store.py` (ledger OOB1, main `92d2ac8`).
# This test reported xpass on the next run, which is the signal to promote it — a
# known_failure left in place after the bug is gone stops guarding anything and starts
# hiding a regression instead.
def test_a_refusal_recorded_by_the_scan_survives_to_the_public_surface():
    """The half of the leak that the band catch-all CANNOT see, by construction.

    `test_no_public_api_response_carries_a_fair_value_past_the_band` walks ratios, and every
    name in this class sits UNDER the band — a refused DCF of 11x is replaced by a peer
    estimate of 3.2x, which is exactly why the catch-all stays green while the leak is open.
    This test asserts the other property: a refusal RECORDED by the scan must still be a
    refusal by the time the public surface reads the row back.

    It uses a real `Store` on a temp file rather than `_FakeStore`, because the defect is in
    persistence and a fake that carries the dict through would prove the opposite of the truth.
    """
    import tempfile

    from valuation.engine.publication import ROW_WITHHELD, record_refusal
    from valuation.screener.fairvalue import estimate_fair_values
    from valuation.screener.store import Store

    rows = _public_rows()
    target = rows[0]
    target["fair_value"] = None
    record_refusal(target, "the model's $1,032.49 is 11.2x the $91.80 price")
    assert target[ROW_WITHHELD] is True, "the fixture did not record a refusal"

    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        st = Store(path)
        st.save_snapshot("2026-08-06", rows, provider="test")
        loaded = st.load_snapshot("2026-08-06")
        back = next(r for r in loaded if r["ticker"] == target["ticker"])

        assert back.get(ROW_WITHHELD), (
            f"the refusal did not survive the snapshot: {target['ticker']} came back with "
            f"{ROW_WITHHELD}={back.get(ROW_WITHHELD)!r}, so the public surface cannot honour it")

        estimate_fair_values(loaded, peer_rows=loaded)
        withhold.withhold_implausible_fair_values(loaded)
        served = next(r for r in loaded if r["ticker"] == target["ticker"])
        assert served.get("fair_value") is None, (
            f"a name the valuation page REFUSES is published at "
            f"{served.get('fair_value')} on the public list")
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_the_public_list_says_withheld_rather_than_going_quiet():
    from valuation.web import app as webapp
    orig = webapp._store
    webapp._store = lambda: _FakeStore(_public_rows())
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            rows = c.get("/api/hotstocks?top=100").get_json()["rows"]
    finally:
        webapp._store = orig
    held = [r for r in rows if r.get("fair_value_withheld")]
    assert len(held) == 2, f"expected AEG and LEV to be withheld, got {len(held)}"
    for r in held:
        assert r["fair_value"] is None and r["upside"] is None
        assert (r.get("fair_value_withheld_reason") or "").strip(), \
            f"{r['ticker']} was blanked with no reason attached"
    assert all(r.get("hot_score") is not None for r in rows), \
        "the RANKING must be untouched — only the fair value is withheld"


def test_the_renderer_shows_a_withheld_row_as_withheld():
    _, src = _render_body()
    i = src.index("function _fairValCell(")
    fn = src[i:src.index("\n}", i)]
    assert "fair_value_withheld" in fn and "withheld" in fn
    assert fn.index("fair_value_withheld") < fn.index("r.fair_value == null"), \
        "a withheld row must not fall through to the missing-data em dash"


def test_the_watchlist_marks_a_partial_score_in_the_cell():
    """`/api/rank` puts a partial score in the same column as full ones. An unmarked 50 beside
    a full 50 asserts they mean the same thing."""
    _, src = _render_body()
    i = src.index("async function runRank(")
    fn = src[i:src.index("\n}\n", i)]
    assert "score_partial" in fn and "PARTIAL" in fn


# --------------------------------------------------------------------------- #
# The exports — a download is a publication
# --------------------------------------------------------------------------- #
# A REAL withheld result, built offline from a synthetic company through the real pipeline,
# so the guard itself decides — no hand-made "withheld" object that could drift from what
# the engine actually produces. NKE's inputs against a $2.00 price come out at 37.5x, which
# is exactly the shape the guard exists for.
def _withheld_result(mc_trials=300):
    from valuation.config import CONFIG
    from valuation.engine.pipeline import value_from_company
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from fixtures import build_nike
    cd = build_nike()
    cd.price = 2.0
    cd.market_cap = 2.0 * cd.shares_diluted
    r = value_from_company(cd, CONFIG, mc_trials=mc_trials)
    assert withhold.is_withheld_result(r), "the fixture stopped tripping the guard"
    return r


def _publishable_result(mc_trials=300):
    from valuation.config import CONFIG
    from valuation.engine.pipeline import value_from_company
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from fixtures import build_nike
    r = value_from_company(build_nike(), CONFIG, mc_trials=mc_trials)
    assert not withhold.is_withheld_result(r)
    return r


def _tmp(suffix):
    import tempfile
    f = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    f.close()
    return f.name


#: Numbers a refusal document may legitimately contain: the price itself, a year, and the
#: figure quoted INSIDE the guard's sentence as the evidence for withholding.
def _implausible_numbers(text, price, reason):
    # collapse whitespace first: a PDF wraps the reason across lines, and an un-normalised
    # comparison would "find" the withheld figure it is allowed to quote
    text = re.sub(r"\s+", " ", str(text))
    text = text.replace(re.sub(r"\s+", " ", reason), " ")   # the one permitted quotation
    # The compute stamp ("2026-08-06 14:32 UTC") is provenance, not money. Removed by an
    # exact shape rather than by loosening the number rule, so nothing shaped like a dollar
    # figure can hide behind it.
    text = re.sub(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2} UTC", " ", text)
    out = []
    for tok in re.findall(r"-?\d[\d,]*\.?\d*", text):
        try:
            v = float(tok.replace(",", ""))
        except ValueError:
            continue
        if 1900 < v < 2200 and "." not in tok:            # a year, not money
            continue
        if v > price * 5:
            out.append(tok)
    return out


def test_the_pdf_renders_the_refusal_instead_of_erroring():
    """A user asking for a tearsheet of a name the model cannot value should get a tearsheet
    that says so, with the reason. An error would claim the export is broken; the export is
    fine and the valuation is withheld, and those are different claims."""
    r = _withheld_result()
    from valuation.report import pdf as pdf_report
    path = _tmp(".pdf")
    pdf_report.build_pdf(r, path)
    assert os.path.getsize(path) > 800, "no document was produced"

    lines = pdf_report.withheld_pdf_lines(r)
    flat = " ".join(str(t[1] if k == "kv" else t) if k != "kv" else f"{t[0]} {t[1]}"
                    for k, t in lines)
    reason = withhold.refusal_reason(r)
    assert reason in flat, "the document must state the reason"
    for phrase in ("not published", "n/a", "not rated"):
        assert phrase in flat, f"the document never says {phrase!r}"
    assert not _implausible_numbers(flat, r.company.price, reason)

    try:
        from pypdf import PdfReader                       # in requirements.txt for this check
    except ImportError:                                   # pragma: no cover
        print("      (pypdf missing — checked the lines, not the rendered file)")
        return
    text = "\n".join((p.extract_text() or "") for p in PdfReader(path).pages)
    assert reason.split(".")[0] in text.replace("\n", " ")
    assert not _implausible_numbers(text, r.company.price, reason), \
        "a figure derived from the withheld valuation is IN THE RENDERED PDF"
    # No SECTION of the normal tearsheet survives. Matched on its headings and table
    # headers, case-sensitively — the refusal's own prose names the withheld cards in
    # lower case ("the bear/base/bull cases ... the Monte Carlo distribution") on purpose,
    # since telling the reader what is missing is the job of the document.
    for banned in ("Score Breakdown", "Reverse DCF", "vs Price", "Base Fair Value",
                   "Comps fair value", "Scenarios"):
        assert banned not in text, f"the refusal tearsheet still has a {banned} section"


def test_the_workbook_has_no_model_in_it():
    """The harder case. A spreadsheet of scenario rows is exactly the shape that leaks, and
    every cell in the normal workbook is a FORMULA — so blanking a summary cell would leave
    the file recomputing the withheld figure the moment it opened. The refusal workbook is
    not the model with holes in it: the model sheets are never built. Checked cell by cell."""
    r = _withheld_result()
    from valuation.report import excel as excel_report
    from openpyxl import load_workbook
    path = _tmp(".xlsx")
    excel_report.build_workbook(r, path)
    wb = load_workbook(path)          # formulas as written, not cached values
    assert wb.sheetnames == ["Not valued"], f"model sheets were built: {wb.sheetnames}"

    reason = withhold.refusal_reason(r)
    cells, formulas = [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cells.append((f"{ws.title}!{c.coordinate}", c.value))
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
    assert formulas == 0, "a refusal workbook must not carry a single live formula"
    assert any(reason in str(v) for _, v in cells), "the reason is not in the workbook"
    for ref, v in cells:
        if isinstance(v, (int, float)):
            assert abs(v) <= r.company.price * 5, f"{ref} = {v} survived"
        else:
            assert not _implausible_numbers(str(v), r.company.price, reason), \
                f"{ref} carries a withheld figure"


def test_a_publishable_name_still_exports_the_whole_model():
    """The refusal must not cost everyone else their workbook."""
    r = _publishable_result()
    from valuation.report import excel as excel_report
    from valuation.report import pdf as pdf_report
    from openpyxl import load_workbook
    xp, pp = _tmp(".xlsx"), _tmp(".pdf")
    excel_report.build_workbook(r, xp)
    pdf_report.build_pdf(r, pp)
    wb = load_workbook(xp)
    assert set(wb.sheetnames) == {"DCF Model", "WACC", "Sensitivity"}
    formulas = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                   if isinstance(c.value, str) and c.value.startswith("="))
    assert formulas > 50, "the live model lost its formulas"
    assert os.path.getsize(pp) > 2000


def test_the_export_routes_serve_the_refusal_document():
    """End to end: 200 and a real file, not a 409. The bytes that reach the browser are
    re-opened and walked, because that is the artefact the user actually gets."""
    from valuation.web import app as webapp
    from openpyxl import load_workbook
    r = _withheld_result()
    reason = withhold.refusal_reason(r)
    webapp._RESULTS.put("NKE", r)
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            pdf = c.get("/api/export/pdf?ticker=NKE")
            assert pdf.status_code == 200, f"pdf returned {pdf.status_code}"
            assert pdf.mimetype == "application/pdf"
            assert pdf.get_data()[:5] == b"%PDF-"

            xls = c.get("/api/export/excel?ticker=NKE")
            assert xls.status_code == 200, f"excel returned {xls.status_code}"
            path = _tmp(".xlsx")
            with open(path, "wb") as fh:
                fh.write(xls.get_data())
            wb = load_workbook(path)
            assert wb.sheetnames == ["Not valued"]
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        assert not str(cell.value).startswith("="), "live formula shipped"
                        assert not _implausible_numbers(str(cell.value),
                                                        r.company.price, reason)
    finally:
        webapp._RESULTS.clear()


def _run_all():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed = failed = xfail = xpass = 0
    for t in tests:
        marked = t.__name__ in _KNOWN_FAILURES
        try:
            t()
        except AssertionError as e:
            if marked:
                xfail += 1
                reason, lane = _KNOWN_FAILURES[t.__name__]
                print(f"  XFAIL {t.__name__}\n         {e}\n         OWNED BY: {lane}")
            else:
                failed += 1
                print(f"  FAIL  {t.__name__}: {e}")
        except Exception as e:
            # A crash is never an XFAIL — a marked test that throws has rotted rather than
            # found something, and filing that under "expected" is how a marker outlives
            # its bug.
            failed += 1
            print(f"  ERROR {t.__name__}: {type(e).__name__}: {e}")
        else:
            if marked:
                xpass += 1
                print(f"  XPASS {t.__name__} — the bug it encodes is FIXED; delete the marker")
            else:
                passed += 1
                print(f"  PASS  {t.__name__}")
    print(f"\n{passed + xfail + xpass}/{len(tests)} withholding tests passed"
          f"  ({xfail} xfail, {xpass} xpass, {failed} failed)")
    if xfail:
        print("XFAIL = a leak this suite can describe but whose repair belongs to another "
              "lane; see HANDOFF_appfixes.md '## BUGS FOUND'.")
    return failed == 0


if __name__ == "__main__":
    sys.exit(0 if _run_all() else 1)
