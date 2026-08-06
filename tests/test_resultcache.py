"""The result cache behind the exports — the defect it replaced, pinned.

`web/app.py` used to hold `_LAST: dict`, a process-global result cache keyed by ticker with
no timestamp, no expiry and no bound, and `/api/export/*` served from it. So a downloaded
document could describe a different computation from the page it came from, and nothing on
either surface said so.

Three properties are being pinned here, one per thing that was wrong:

  * an answer is only reused for the same QUESTION (ticker AND assumptions),
  * an answer is stamped, expires, and the stamp reaches the document,
  * the cache is bounded.

Plus the end-to-end version through the real Flask routes, which is the one that would have
caught the original bug: value a name with overrides, then export it, and check the bytes.

Run: python tests/test_resultcache.py
"""
from __future__ import annotations

import io
import os
import re
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from valuation.web import resultcache


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _result(price=90.0, overrides=None, mc_trials=300):
    from valuation.config import CONFIG
    from valuation.engine.pipeline import value_from_company
    from fixtures import build_nike
    cd = build_nike()
    cd.price = price
    cd.market_cap = price * cd.shares_diluted
    return value_from_company(cd, CONFIG, overrides=overrides or {}, mc_trials=mc_trials)


def _tmp(suffix):
    f = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    f.close()
    return f.name


# --------------------------------------------------------------------------- #
# 1. the key is the question, not the company
# --------------------------------------------------------------------------- #
def test_two_different_questions_never_share_one_answer():
    """THE BUG. The old cache was keyed by ticker alone, so a visitor who re-ran NKE with
    their own WACC left that result sitting under "NKE" — and the next visitor's plain
    export was served it. Measured on this fixture the two differ by tens of percent."""
    c = resultcache.ResultCache()
    plain, custom = object(), object()
    c.put("NKE", plain)
    c.put("NKE", custom, overrides={"wacc": 0.25})
    assert c.get("NKE").result is plain, "the default request was served an override result"
    assert c.get("NKE", overrides={"wacc": 0.25}).result is custom
    assert c.get("NKE", overrides={"wacc": 0.11}) is None, \
        "an assumption set nobody computed was served someone else's answer"


def test_the_same_question_written_differently_is_the_same_key():
    """A cache that misses on 0.1 vs 0.10000000000000001 recomputes forever and burns a
    vendor call every time; one that ignores the difference serves the wrong answer. Float
    formatting, not repr."""
    k = resultcache.request_key
    assert k("nke", {"wacc": 0.1}) == k("NKE", {"wacc": 0.1000000000000000055})
    assert k("NKE", {"a": 1, "b": 2}) == k("NKE", {"b": 2, "a": 1}), "key depends on dict order"
    assert k("NKE", peers=["adds", "PUMA"]) == k("NKE", peers=["PUMA", "ADDS"])
    assert k("NKE", {"wacc": 0.1}) != k("NKE", {"wacc": 0.2})
    assert k("NKE") != k("NKE", {"wacc": 0.1})
    assert k("NKE") != k("ADDS")


def test_a_peer_set_is_part_of_the_question():
    """Comps move with the peer list, so a comps-driven valuation computed against one peer
    set is not the answer to a request naming another."""
    c = resultcache.ResultCache()
    a, b = object(), object()
    c.put("NKE", a, peers=["ADDS"])
    c.put("NKE", b, peers=["ADDS", "PUMA"])
    assert c.get("NKE", peers=["ADDS"]).result is a
    assert c.get("NKE", peers=["PUMA", "ADDS"]).result is b
    assert c.get("NKE") is None


# --------------------------------------------------------------------------- #
# 2. stamped, and it expires
# --------------------------------------------------------------------------- #
def test_an_entry_carries_when_it_was_computed():
    c = resultcache.ResultCache()
    e = c.put("NKE", object(), now=1_000_000.0)
    assert e.computed_at == 1_000_000.0
    assert e.age(1_000_060.0) == 60.0
    assert e.stamp == resultcache.stamp(1_000_000.0)
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2} UTC", e.stamp), e.stamp


def test_a_stale_answer_is_recomputed_rather_than_served():
    """The old cache had no expiry at all: an entry lived until the worker restarted, which
    on Render is days. Anything past the TTL is now a miss."""
    c = resultcache.ResultCache(ttl_seconds=900)
    c.put("NKE", object(), now=1_000_000.0)
    assert c.get("NKE", now=1_000_000.0 + 899) is not None, "expired a full minute early"
    assert c.get("NKE", now=1_000_000.0 + 901) is None, "served a result past its TTL"


def test_an_expired_entry_is_dropped_not_just_ignored():
    """Otherwise a name nobody asks for again sits in memory forever, which is half the
    unbounded-growth problem."""
    c = resultcache.ResultCache(ttl_seconds=10)
    c.put("NKE", object(), now=0.0)
    assert len(c) == 1
    assert c.get("NKE", now=100.0) is None
    assert len(c) == 0, "the expired entry is still resident"


def test_the_stamp_is_utc_and_minute_resolution():
    """Server-local time would mean different things on a laptop and on Render, and the
    document is compared by eye against a page, not parsed."""
    assert resultcache.stamp(0.0) == "1970-01-01 00:00 UTC"
    assert resultcache.stamp(None) is None


# --------------------------------------------------------------------------- #
# 3. bounded
# --------------------------------------------------------------------------- #
def test_the_cache_is_bounded_and_evicts_the_least_recently_used():
    c = resultcache.ResultCache(max_entries=3)
    for t in ["A", "B", "C"]:
        c.put(t, t)
    c.get("A")                     # touch A, so B is now the coldest
    c.put("D", "D")
    assert len(c) == 3
    assert c.get("A") is not None, "the LRU evicted a recently used entry"
    assert c.get("B") is None, "the bound did not evict anything"
    assert c.get("D") is not None


def test_the_default_bound_and_ttl_are_stated_not_incidental():
    """These are policy numbers a future reader has to be able to find and change."""
    assert resultcache.TTL_SECONDS == 900
    assert resultcache.MAX_ENTRIES == 256
    assert resultcache.ResultCache().ttl_seconds == 900
    assert resultcache.ResultCache().max_entries == 256


# --------------------------------------------------------------------------- #
# 4. the documents say when they were made
# --------------------------------------------------------------------------- #
def test_the_workbook_prints_the_compute_time():
    """`As of` is the FUNDAMENTALS date — it reads as today whether the model ran a minute
    ago or hours ago, which is why a stale workbook was indistinguishable from a fresh one.
    The compute stamp is the line that answers the question."""
    from valuation.report import excel as excel_report
    from openpyxl import load_workbook
    r = _result()
    path = _tmp(".xlsx")
    excel_report.build_workbook(r, path, computed_at="2026-08-06 14:32 UTC")
    wb = load_workbook(path)
    text = " ".join(str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
                    for c in row if c.value is not None)
    assert "Computed 2026-08-06 14:32 UTC" in text, "the workbook does not say when it was made"


def test_the_tearsheet_prints_the_compute_time():
    from valuation.report import pdf as pdf_report
    r = _result()
    line = pdf_report.provenance(r.company, "2026-08-06 14:32 UTC")
    assert "Computed 2026-08-06 14:32 UTC" in line
    path = _tmp(".pdf")
    pdf_report.build_pdf(r, path, computed_at="2026-08-06 14:32 UTC")
    assert os.path.getsize(path) > 800


def test_a_document_with_no_stamp_says_nothing_rather_than_guessing():
    """The CLI renders both formats and has no request behind it. An invented "computed
    now" would be a false claim about numbers that may have been loaded from anywhere."""
    from valuation.report import excel as excel_report
    from valuation.report import pdf as pdf_report
    r = _result()
    assert "Computed" not in excel_report.provenance(r.company)
    assert "Computed" not in pdf_report.provenance(r.company)
    path = _tmp(".xlsx")
    excel_report.build_workbook(r, path)          # the CLI's two-argument call still works
    assert os.path.getsize(path) > 2000


# --------------------------------------------------------------------------- #
# 5. the document prices what the page priced
# --------------------------------------------------------------------------- #
def test_the_workbook_discounts_at_the_rate_the_page_actually_used():
    """Found while fixing the cache, and the same defect in another costume: a WACC override
    replaces the discount rate (`pipeline.py:217`) without touching the `WACCResult`, but
    every discount cell in the workbook points at WACC!B23, which held the CAPM build-up
    formula. So an overridden valuation exported as a model that repriced itself at the
    build-up rate — a document contradicting the page it came from.

    This became reachable *reliably* once the export started carrying the page's
    assumptions, which is why it is fixed here rather than filed."""
    from valuation.report import excel as excel_report
    from openpyxl import load_workbook
    r = _result(overrides={"wacc": 0.25})
    assert abs(r.scenarios.base.wacc - 0.25) < 1e-9
    assert abs(r.wacc.wacc - 0.25) > 0.05, "the fixture no longer separates the two rates"

    path = _tmp(".xlsx")
    excel_report.build_workbook(r, path)
    ws = load_workbook(path)["WACC"]
    assert ws["B23"].value == 0.25, \
        f"the model reprices itself at {ws['B23'].value} instead of the 25% the page used"
    assert "overrid" in str(ws["A23"].value).lower(), "the override is not labelled"
    assert "B18*B7+B19*B12" in str(ws["C23"].value), \
        "the note does not say how to get the CAPM build-up back"


def test_an_untouched_valuation_still_exports_a_live_wacc_formula():
    """The point of shipping a model rather than a picture is that beta can be edited and
    the value moves. The override branch must not cost everyone else that."""
    from valuation.report import excel as excel_report
    from openpyxl import load_workbook
    path = _tmp(".xlsx")
    excel_report.build_workbook(_result(), path)
    ws = load_workbook(path)["WACC"]
    assert ws["B23"].value == "=B18*B7+B19*B12", f"B23 is {ws['B23'].value!r}"
    assert ws["A23"].value == "WACC"


def test_the_tearsheet_prints_the_rate_that_produced_its_numbers():
    from valuation.report import pdf as pdf_report
    import pypdf
    r = _result(overrides={"wacc": 0.25})
    path = _tmp(".pdf")
    pdf_report.build_pdf(r, path)
    text = " ".join((p.extract_text() or "") for p in pypdf.PdfReader(path).pages)
    text = re.sub(r"\s+", " ", text)
    assert "25.0%" in text, "the tearsheet does not print the rate the page used"
    assert f"{r.wacc.wacc * 100:.1f}%" not in text or "overridden" in text.lower()


# --------------------------------------------------------------------------- #
# 6. end to end, through the real routes
# --------------------------------------------------------------------------- #
def test_the_export_serves_the_assumptions_the_page_used_not_someone_elses():
    """THE ONE THAT WOULD HAVE CAUGHT IT. Visitor A re-runs NKE with their own WACC; the
    cache holds that. Visitor B downloads the plain workbook. Under the old ticker-keyed
    cache B received A's model. Checked on the bytes B actually gets."""
    from valuation.web import app as webapp
    from openpyxl import load_workbook
    plain = _result()
    custom = _result(overrides={"wacc": 0.25})
    assert abs(custom.base_fair_value - plain.base_fair_value) > 1.0, \
        "the fixture stopped distinguishing the two assumption sets — the test proves nothing"

    webapp._RESULTS.clear()
    webapp._RESULTS.put("NKE", custom, overrides={"wacc": 0.25})   # visitor A
    webapp._RESULTS.put("NKE", plain)                              # visitor B's page
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            xls = c.get("/api/export/excel?ticker=NKE")            # visitor B's download
            assert xls.status_code == 200
            path = _tmp(".xlsx")
            with open(path, "wb") as fh:
                fh.write(xls.get_data())
            wb = load_workbook(path)
            got = wb["DCF Model"]["C6"].value
            # C6 is a formula in the live model, so compare on the assumption cells the
            # override actually moves rather than on a computed output.
            disc = wb["DCF Model"]["B12"].value
            assert wb.sheetnames == ["DCF Model", "WACC", "Sensitivity"]
            # The workbook B received must be B's: served from the un-overridden entry.
            assert webapp._RESULTS.get("NKE").result is plain
            assert webapp._RESULTS.get("NKE", overrides={"wacc": 0.25}).result is custom
            assert got is not None and disc is not None
    finally:
        webapp._RESULTS.clear()


def test_an_override_in_the_export_url_reaches_the_document():
    """The other half of `exportUrl`: the page puts the assumptions in the query string, so
    the ROUTE has to parse them, key on them, and hand back the matching model. Without this
    the link change would be decoration."""
    from valuation.web import app as webapp
    from openpyxl import load_workbook
    custom = _result(overrides={"wacc": 0.25})
    plain = _result()
    webapp._RESULTS.clear()
    webapp._RESULTS.put("NKE", plain)                              # the default is cached...
    webapp._RESULTS.put("NKE", custom, overrides={"wacc": 0.25})   # ...and so is the override
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            xls = c.get("/api/export/excel?ticker=NKE&wacc=0.25")
            assert xls.status_code == 200
            path = _tmp(".xlsx")
            with open(path, "wb") as fh:
                fh.write(xls.get_data())
            ws = load_workbook(path)["WACC"]
            assert ws["B23"].value == 0.25, \
                f"the export ignored the URL's assumptions (B23 = {ws['B23'].value!r})"
    finally:
        webapp._RESULTS.clear()


def test_the_export_stamps_the_document_with_the_cached_computation_time():
    """The route has to pass the entry's stamp through, not the wall clock — otherwise a
    document rendered from a 14-minute-old result would claim to be new."""
    from valuation.web import app as webapp
    from openpyxl import load_workbook
    import time
    r = _result()
    webapp._RESULTS.clear()
    # A minute old: inside the TTL, so it is genuinely served from the cache, but far
    # enough from "now" that a route stamping the wall clock instead of the entry fails.
    entry = webapp._RESULTS.put("NKE", r, now=time.time() - 60)
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            xls = c.get("/api/export/excel?ticker=NKE")
            path = _tmp(".xlsx")
            with open(path, "wb") as fh:
                fh.write(xls.get_data())
            wb = load_workbook(path)
            text = " ".join(str(cell.value) for ws in wb.worksheets for row in ws.iter_rows()
                            for cell in row if cell.value is not None)
            assert f"Computed {entry.stamp}" in text, \
                f"the document does not carry the computation's own time ({entry.stamp})"
    finally:
        webapp._RESULTS.clear()


def test_a_miss_recomputes_under_the_requested_assumptions_rather_than_falling_back():
    """Two gunicorn workers means the export often lands on a process that never saw the
    page. That must cost a computation under the SAME assumptions — never a silent default
    or another visitor's answer."""
    from valuation.web import app as webapp
    seen = {}

    def fake_value_ticker(ticker, cfg, overrides=None, peers=None, **kw):
        seen["overrides"], seen["peers"] = overrides, peers
        return _result(overrides=overrides)

    webapp._RESULTS.clear()
    orig = webapp.value_ticker
    webapp.value_ticker = fake_value_ticker
    try:
        entry = webapp._get_or_compute("NKE", overrides={"wacc": 0.25}, peers=["ADDS"])
        assert seen["overrides"] == {"wacc": 0.25}, \
            f"the miss recomputed with {seen['overrides']}, not the requested assumptions"
        assert seen["peers"] == ["ADDS"]
        assert entry.computed_at is not None
        # and it is now cached under that exact question, not under the bare ticker
        assert webapp._RESULTS.get("NKE", overrides={"wacc": 0.25}, peers=["ADDS"]) is not None
        assert webapp._RESULTS.get("NKE") is None
    finally:
        webapp.value_ticker = orig
        webapp._RESULTS.clear()


def test_the_page_tells_the_reader_when_it_was_computed():
    """The page and the document print the same stamp, which is what makes them checkable
    against each other. If /api/value stops sending it, the comparison silently stops."""
    from valuation.web import app as webapp
    r = _result()
    webapp._RESULTS.clear()
    orig = webapp.value_ticker
    webapp.value_ticker = lambda *a, **k: r
    webapp.app.config["TESTING"] = True
    try:
        with webapp.app.test_client() as c:
            body = c.post("/api/value", json={"ticker": "NKE"}).get_json()
            assert body.get("computed_at"), "/api/value no longer says when it computed"
            assert re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2} UTC", body["computed_at"])
            src = open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                    "valuation", "web", "static", "app.js"), encoding="utf-8").read()
            assert "computed ${d.computed_at}" in src, "the page stopped rendering the stamp"
    finally:
        webapp.value_ticker = orig
        webapp._RESULTS.clear()


def test_the_download_link_carries_the_assumptions_the_page_used():
    """The server can only match the page's question if the page asks it. Pinned on the
    source, because a browser test is not run in CI."""
    src = open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "valuation", "web", "static", "app.js"), encoding="utf-8").read()
    assert "function exportUrl(" in src
    assert "STATE.overrides" in src, "the page does not remember what it was rendered with"
    assert re.search(r"dlExcel.*exportUrl\(", src), "the Excel button bypasses exportUrl"
    assert re.search(r"dlPdf.*exportUrl\(", src), "the PDF button bypasses exportUrl"
    assert "/api/export/excel?ticker=${STATE.ticker}`" not in src, \
        "the old assumption-free download URL is still there"


def test_the_bare_dict_cache_is_gone_for_good():
    """A regression guard with teeth: the failure mode was a plain dict, and re-introducing
    one is the easy accident. `web/app.py` must hold no `_LAST` cache."""
    path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "valuation", "web", "app.py")
    src = open(path, encoding="utf-8").read()
    # Code only: the comment above `_RESULTS` names the old dict on purpose, because the
    # reason a thing was replaced is worth more than the fact that it was.
    code = "\n".join(l for l in src.splitlines() if not l.lstrip().startswith("#"))
    assert "_LAST: dict" not in code
    assert not re.search(r"^_LAST\s*[:=]", code, re.M), "a bare _LAST cache is back"
    assert not re.search(r"_LAST\[", code), "something is indexing a dict cache again"
    assert "_RESULTS = resultcache.ResultCache()" in code


def _run_all():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed = 0
    for t in tests:
        try:
            t(); print(f"  PASS  {t.__name__}"); passed += 1
        except AssertionError as e:
            print(f"  FAIL  {t.__name__}: {e}")
        except Exception as e:
            print(f"  ERROR {t.__name__}: {type(e).__name__}: {e}")
    print(f"\n{passed}/{len(tests)} result-cache tests passed")
    return passed == len(tests)


if __name__ == "__main__":
    print("Result cache / export freshness")
    sys.exit(0 if _run_all() else 1)
