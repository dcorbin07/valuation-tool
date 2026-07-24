"""
Excel export — a live, formula-driven DCF workbook for any ticker.

Deliberately mirrors the layout of Donovan's original Nike model: a DCF Model
sheet, a WACC sheet, and a Sensitivity sheet, all driven by formulas that
recalculate when you change an input. Follows financial-model conventions:
blue text for hardcoded inputs, black for formulas, green for cross-sheet links;
currency/percent/multiple number formats; Arial throughout.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", color="0000FF", size=10)       # hardcoded inputs
BLACK = Font(name="Arial", color="000000", size=10)       # formulas
GREEN = Font(name="Arial", color="008000", size=10)       # cross-sheet links
BOLD = Font(name="Arial", bold=True, size=10)
TITLE = Font(name="Arial", bold=True, size=13)
HDR = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SECT_FILL = PatternFill("solid", fgColor="D9E1F2")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);-'
CUR2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'
MULT = '0.0x'


def _c(ws, ref, value, font=BLACK, fmt=None, fill=None, align=None, border=False):
    cell = ws[ref]
    cell.value = value
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = Alignment(horizontal=align)
    if border:
        cell.border = BORDER
    return cell


def build_workbook(result, path: str) -> str:
    cd = result.company
    a = result.assumptions
    w = result.wacc
    base = result.scenarios.base
    n = a.n_years

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for i in range(2, 3 + n + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # Columns: B = base year (actual), C.. = forecast years 1..n
    base_col = 2
    first = 3
    last = 3 + n - 1
    L = lambda i: get_column_letter(i)
    lastL, firstL, baseL = L(last), L(first), L(base_col)

    # ---- Title ----
    _c(ws, "A1", f"{cd.name} ({cd.ticker}) — Discounted Cash Flow Valuation", TITLE)
    _c(ws, "A2", f"Regime: {result.classification.regime} | Currency: {cd.currency} | "
                 f"$ in millions | As of {cd.as_of}", Font(name="Arial", italic=True, size=9))
    _c(ws, "A3", f"Unlevered FCFF / reinvestment method | Reference price: "
                 f"${cd.price:,.2f}" if cd.price else "Reference price: n/a",
       Font(name="Arial", italic=True, size=9))

    # ---- Output summary ----
    _c(ws, "A5", "OUTPUT SUMMARY", BOLD, fill=SECT_FILL)
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws[f"{col}5"].fill = SECT_FILL
    _c(ws, "A6", "Implied value per share")
    _c(ws, "C6", f"={baseL}41", BOLD, CUR2)
    _c(ws, "E6", "Enterprise value ($mm)")
    _c(ws, "G6", f"={baseL}36", BLACK, CUR)
    _c(ws, "A7", "Current reference price")
    _c(ws, "C7", cd.price if cd.price else 0, BLUE, CUR2)
    _c(ws, "E7", "Upside / (downside)")
    _c(ws, "G7", "=C6/C7-1", BLACK, PCT)
    _c(ws, "A8", "WACC")
    _c(ws, "C8", "=WACC!B23", GREEN, PCT)
    _c(ws, "E8", "% of EV from terminal value")
    _c(ws, "G8", f"={baseL}35/{baseL}36", BLACK, PCT)
    _c(ws, "A9", "Opportunity score (1-100)")
    _c(ws, "C9", f"{result.score.score}  ({result.score.recommendation})", BOLD)

    # ---- Key assumptions ----
    _c(ws, "A11", "KEY ASSUMPTIONS", BOLD, fill=SECT_FILL)
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}11"].fill = SECT_FILL
    _c(ws, "A12", "Normalized tax rate")
    _c(ws, "B12", a.tax_rate, BLUE, PCT, fill=YELLOW)
    _c(ws, "D12", "Terminal growth rate")
    _c(ws, "E12", a.terminal_growth, BLUE, PCT, fill=YELLOW)
    _c(ws, "A13", "Sales-to-capital (reinvestment)")
    _c(ws, "B13", a.sales_to_capital, BLUE, MULT, fill=YELLOW)
    _c(ws, "D13", "Terminal ROIC")
    _c(ws, "E13", base.terminal_roic, BLUE, PCT, fill=YELLOW)
    _c(ws, "A14", "Diluted shares (mm)")
    _c(ws, "B14", cd.shares_diluted if cd.shares_diluted else 0, BLUE, '#,##0')
    _c(ws, "D14", "Net debt ($mm)")
    _c(ws, "E14", cd.net_debt if cd.net_debt is not None else 0, BLUE, CUR)

    # ---- Projection ----
    r = 16
    _c(ws, f"A{r}", "FREE CASH FLOW PROJECTION", BOLD, fill=SECT_FILL)
    for i in range(2, last + 1):
        ws[f"{L(i)}{r}"].fill = SECT_FILL
    # header row
    rH = r + 1
    _c(ws, f"A{rH}", "Fiscal year", HDR, fill=HDR_FILL)
    fy0 = cd.fiscal_years[0] if cd.fiscal_years else "FY0"
    _c(ws, f"{baseL}{rH}", f"{fy0} (A)", HDR, fill=HDR_FILL, align="center")
    for k in range(n):
        _c(ws, f"{L(first+k)}{rH}", f"Yr {k+1}", HDR, fill=HDR_FILL, align="center")

    rG = rH + 1     # revenue growth
    rRev = rG + 1   # revenue
    rM = rRev + 1   # margin
    rE = rM + 1     # EBIT
    rT = rE + 1     # taxes
    rNO = rT + 1    # NOPAT
    rRI = rNO + 1   # reinvestment
    rF = rRI + 1    # FCFF
    rN = rF + 1     # period
    rD = rN + 1     # discount factor
    rPV = rD + 1    # PV

    _c(ws, f"A{rG}", "Revenue growth %")
    _c(ws, f"{baseL}{rG}", "—", BLACK, align="center")
    _c(ws, f"A{rRev}", "Revenue", BOLD)
    _c(ws, f"{baseL}{rRev}", a.base_revenue, BLUE, CUR)
    _c(ws, f"A{rM}", "Operating (EBIT) margin %")
    _c(ws, f"{baseL}{rM}", a.current_margin, BLUE, PCT)
    _c(ws, f"A{rE}", "EBIT")
    _c(ws, f"{baseL}{rE}", f"={baseL}{rRev}*{baseL}{rM}", BLACK, CUR)
    _c(ws, f"A{rT}", "Less: cash taxes on EBIT")
    _c(ws, f"A{rNO}", "NOPAT")
    _c(ws, f"A{rRI}", "Less: reinvestment (ΔRev ÷ S2C)")
    _c(ws, f"A{rF}", "Unlevered FCFF", BOLD)
    _c(ws, f"A{rN}", "Period (n)")
    _c(ws, f"A{rD}", "Discount factor @ WACC")
    _c(ws, f"A{rPV}", "PV of FCFF", BOLD)

    for k in range(n):
        col = L(first + k)
        prev = baseL if k == 0 else L(first + k - 1)
        _c(ws, f"{col}{rG}", a.rev_growth_path[k], BLUE, PCT, align="center")
        _c(ws, f"{col}{rRev}", f"={prev}{rRev}*(1+{col}{rG})", BLACK, CUR)
        _c(ws, f"{col}{rM}", a.op_margin_path[k], BLUE, PCT, align="center")
        _c(ws, f"{col}{rE}", f"={col}{rRev}*{col}{rM}", BLACK, CUR)
        _c(ws, f"{col}{rT}", f"=-IF({col}{rE}>0,{col}{rE}*$B${12},0)", BLACK, CUR)
        _c(ws, f"{col}{rNO}", f"={col}{rE}+{col}{rT}", BLACK, CUR)
        _c(ws, f"{col}{rRI}", f"=-({col}{rRev}-{prev}{rRev})/$B${13}", BLACK, CUR)
        _c(ws, f"{col}{rF}", f"={col}{rNO}+{col}{rRI}", BOLD, CUR)
        _c(ws, f"{col}{rN}", k + 1, BLACK, align="center")
        _c(ws, f"{col}{rD}", f"=1/(1+WACC!$B$23)^{col}{rN}", BLACK, '0.000')
        _c(ws, f"{col}{rPV}", f"={col}{rF}*{col}{rD}", BLACK, CUR)

    # ---- Valuation bridge (single values live in column C, like the Nike model) ----
    valL = "C"
    rB = rPV + 2
    _c(ws, f"A{rB}", "VALUATION BRIDGE", BOLD, fill=SECT_FILL)
    for i in range(2, last + 1):
        ws[f"{L(i)}{rB}"].fill = SECT_FILL
    rows = {
        "pv_expl": rB + 1, "tv": rB + 2, "pv_tv": rB + 3, "ev": rB + 4,
        "nd": rB + 5, "eq": rB + 6, "sh": rB + 7, "ps": rB + 8, "px": rB + 9, "up": rB + 10,
    }
    _c(ws, f"A{rows['pv_expl']}", "PV of explicit FCFF")
    _c(ws, f"{valL}{rows['pv_expl']}", f"=SUM({firstL}{rPV}:{lastL}{rPV})", BLACK, CUR)
    _c(ws, f"A{rows['tv']}", "Terminal value (Gordon, ROIC-consistent)")
    # TV = last_rev*(1+g)*term_margin*(1-tax)*(1 - g/ROIC) / (WACC - g)
    tv_formula = (f"={lastL}{rRev}*(1+$E${12})*{lastL}{rM}*(1-$B${12})*"
                  f"(1-$E${12}/$E${13})/(WACC!$B$23-$E${12})")
    _c(ws, f"{valL}{rows['tv']}", tv_formula, BLACK, CUR)
    _c(ws, f"A{rows['pv_tv']}", "PV of terminal value")
    _c(ws, f"{valL}{rows['pv_tv']}", f"={valL}{rows['tv']}*{lastL}{rD}", BLACK, CUR)
    _c(ws, f"A{rows['ev']}", "Enterprise value", BOLD)
    _c(ws, f"{valL}{rows['ev']}", f"={valL}{rows['pv_expl']}+{valL}{rows['pv_tv']}", BOLD, CUR)
    _c(ws, f"A{rows['nd']}", "Less: net debt")
    _c(ws, f"{valL}{rows['nd']}", f"=-$E${14}", BLACK, CUR)
    _c(ws, f"A{rows['eq']}", "Equity value", BOLD)
    _c(ws, f"{valL}{rows['eq']}", f"={valL}{rows['ev']}+{valL}{rows['nd']}", BOLD, CUR)
    _c(ws, f"A{rows['sh']}", "Diluted shares (mm)")
    _c(ws, f"{valL}{rows['sh']}", f"=$B${14}", BLACK, '#,##0')
    _c(ws, f"A{rows['ps']}", "Implied value per share", BOLD)
    _c(ws, f"{valL}{rows['ps']}", f"={valL}{rows['eq']}/{valL}{rows['sh']}", BOLD, CUR2)
    _c(ws, f"A{rows['px']}", "Current reference price")
    _c(ws, f"{valL}{rows['px']}", "=C7", BLACK, CUR2)
    _c(ws, f"A{rows['up']}", "Implied upside / (downside)")
    _c(ws, f"{valL}{rows['up']}", f"={valL}{rows['ps']}/{valL}{rows['px']}-1", BLACK, PCT)

    # Point the output-summary cells at the actual bridge rows.
    ws["C6"].value = f"={valL}{rows['ps']}"
    ws["G6"].value = f"={valL}{rows['ev']}"
    ws["G8"].value = f"={valL}{rows['pv_tv']}/{valL}{rows['ev']}"

    _build_wacc_sheet(wb, cd, w)
    _build_sensitivity_sheet(wb, cd, a, base, rRev, rM, rF, rN, first, last, n)

    wb.save(path)
    return path


def _build_wacc_sheet(wb, cd, w):
    ws = wb.create_sheet("WACC")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 44
    _c(ws, "A1", f"WACC — {cd.name}", TITLE)
    _c(ws, "A3", "Cost of Equity (CAPM)", BOLD, fill=SECT_FILL); ws["B3"].fill = SECT_FILL; ws["C3"].fill = SECT_FILL
    _c(ws, "A4", "Risk-free rate (10Y UST)"); _c(ws, "B4", w.risk_free, BLUE, PCT, fill=YELLOW)
    _c(ws, "C4", "Live 10Y Treasury at run time.", Font(name="Arial", italic=True, size=9))
    _c(ws, "A5", "Equity risk premium"); _c(ws, "B5", w.erp, BLUE, PCT, fill=YELLOW)
    _c(ws, "C5", "Default ~5.0% (Damodaran).", Font(name="Arial", italic=True, size=9))
    _c(ws, "A6", "Levered beta"); _c(ws, "B6", w.beta, BLUE, '0.00', fill=YELLOW)
    _c(ws, "A7", "Cost of equity", BOLD); _c(ws, "B7", "=B4+B6*B5", BOLD, PCT)
    _c(ws, "A9", "Cost of Debt", BOLD, fill=SECT_FILL); ws["B9"].fill = SECT_FILL; ws["C9"].fill = SECT_FILL
    _c(ws, "A10", "Pre-tax cost of debt"); _c(ws, "B10", w.cost_of_debt_pretax, BLUE, PCT, fill=YELLOW)
    _c(ws, "A11", "Tax rate"); _c(ws, "B11", "='DCF Model'!B12", GREEN, PCT)
    _c(ws, "A12", "After-tax cost of debt", BOLD); _c(ws, "B12", "=B10*(1-B11)", BOLD, PCT)
    _c(ws, "A14", "Capital Structure ($mm, market)", BOLD, fill=SECT_FILL); ws["B14"].fill = SECT_FILL; ws["C14"].fill = SECT_FILL
    _c(ws, "A15", "Market value of equity"); _c(ws, "B15", w.market_value_equity, BLUE, CUR)
    _c(ws, "A16", "Market value of debt"); _c(ws, "B16", w.market_value_debt, BLUE, CUR)
    _c(ws, "A17", "Total capital"); _c(ws, "B17", "=B15+B16", BLACK, CUR)
    _c(ws, "A18", "Weight of equity"); _c(ws, "B18", "=B15/B17", BLACK, PCT)
    _c(ws, "A19", "Weight of debt"); _c(ws, "B19", "=B16/B17", BLACK, PCT)
    _c(ws, "A23", "WACC", BOLD); _c(ws, "B23", "=B18*B7+B19*B12", BOLD, PCT, fill=YELLOW)


def _build_sensitivity_sheet(wb, cd, a, base, rRev, rM, rF, rN, first, last, n):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws.column_dimensions[col].width = 12
    _c(ws, "A1", "Sensitivity — Implied Share Price", TITLE)
    _c(ws, "A2", "Rows: WACC   |   Columns: terminal growth", Font(name="Arial", italic=True, size=9))
    L = get_column_letter
    firstL, lastL = L(first), L(last)
    dm = "'DCF Model'!"
    fcff_rng = f"{dm}${firstL}${rF}:${lastL}${rF}"
    per_rng = f"{dm}${firstL}${rN}:${lastL}${rN}"
    last_rev = f"{dm}${lastL}${rRev}"
    term_m = f"{dm}${lastL}${rM}"
    tax = f"{dm}$B$12"
    troic = f"{dm}$E$13"
    nd = f"{dm}$E$14"
    sh = f"{dm}$B$14"

    _c(ws, "A4", "WACC \\ g", BOLD, align="center", fill=SECT_FILL)
    g_base = base.terminal_growth
    w_base = base.wacc
    gcols = ["B", "C", "D", "E", "F"]
    for j, col in enumerate(gcols):
        gval = round(g_base + (j - 2) * 0.005, 4)
        _c(ws, f"{col}4", gval, BLUE, PCT, align="center", fill=SECT_FILL)
    for i in range(5):
        row = 5 + i
        wval = round(w_base + (i - 2) * 0.01, 4)
        _c(ws, f"A{row}", wval, BLUE, PCT, align="center", fill=SECT_FILL)
        for col in gcols:
            wref = f"$A{row}"
            gref = f"{col}$4"
            # per_share = (SUMPRODUCT(FCFF,(1+w)^-n) + TV*(1+w)^-N - net_debt)/shares
            tv = (f"({last_rev}*(1+{gref})*{term_m}*(1-{tax})*(1-{gref}/{troic})/({wref}-{gref}))")
            disc_n = f"(1+{wref})^(-{dm}${lastL}${rN})"
            formula = (f"=(SUMPRODUCT({fcff_rng},(1+{wref})^(-{per_rng}))+{tv}*{disc_n}-{nd})/{sh}")
            _c(ws, f"{col}{row}", formula, BLACK, CUR2, border=True)
    _c(ws, "A11", "Base case shaded conceptually at center (WACC≈base, g≈base).",
       Font(name="Arial", italic=True, size=9))
    # highlight center cell
    ws["D7"].fill = YELLOW
    return ws
